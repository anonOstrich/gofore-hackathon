from openpyxl import load_workbook

def simple():
    return str(200)

def get_name_info_xlsx(name="Tomi"):
    book = load_workbook('data/etunimitilasto-2019-01-07-vrk.xlsx')
    count = 0
    idxs = [1, 4]
    for sheet_index in idxs:
        sheet = book.worksheets[sheet_index]
        for idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_row=10000, max_col=1)):
            if row[0].value == None:
                break
            for cell in row:
                if cell.value.lower() == name.lower():
                    count = max(count, sheet['B' + str(idx+1)].value)
    return str(count)

def name_xlsx_to_dict(filename):
    dict = {}
    book = load_workbook(filename)
    idxs = [1,4]
    for sheet_index in idxs:
        sheet = book.worksheets[sheet_index]
        for row in sheet.iter_rows(min_row=2, min_col=1, max_row=10000, max_col=2):
            if row[0].value == None:
                break
            this_name = row[0].value.lower()
            this_count = row[1].value
            if dict.get(this_name) != None:
                if dict[this_name] < this_count:
                    dict[this_name] = this_count
            else:
                dict[this_name] = this_count

    return dict
